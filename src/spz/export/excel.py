# -*- coding: utf-8 -*-

"""Formatter that writes excel files.
"""

from . import TableWriter

import re

from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl.worksheet.table import Table
from openpyxl.workbook.child import INVALID_TITLE_REGEX
from zipfile import ZipFile
from spz import app


def find_table(workbook, table_name):
    for sheet in workbook.worksheets:
        for table in sheet._tables:
            return sheet, table

    """function to find information table with the course information
    in table: Level, ECTS, course name, date of exam
    """


def find_course_table(name, tables):
    for table in tables:
        if table.displayName == name:
            return table.ref


def delete_last_row(sheet, range):
    for c in range.bottom:
        sheet.cell(*c).value = None
    range.shrink(bottom=1)


def sanitize_title(str):
    """Sanitize a string so it can safely be used as a worksheet title:
    Disallowed characters will be stripped and the string will be trimmed to a length of 30.
    """
    return re.sub(INVALID_TITLE_REGEX, '', str)[:30]


class ExcelWriter(TableWriter):
    """The base ExcelWriter begins a new sheet for each new section.
    """

    @property
    def mimetype(self):
        return self.workbook.mime_type

    @property
    def extension(self):
        return 'xltx' if self.workbook.template else 'xlsx'

    def __init__(self, template):
        TableWriter.__init__(self, template, binary_template=True)
        self.section_count = 0

    def parse_template(self, file):
        self.workbook = load_workbook(file)
        # once we have the template sheet and its index, we remove it from the workbook
        # this way we can add multiple copies of it while keeping the original unmodified
        self.template_sheet, self.template_table = find_table(self.workbook, 'DATA')
        self.sheet_insert_index = self.workbook.index(self.template_sheet)
        self.template_range = CellRange(self.template_table.ref)
        expression_row = [self.template_sheet.cell(*c).value for c in self.template_range.bottom]
        delete_last_row(self.template_sheet, self.template_range)  # this row contains the jinja-expressions
        self.workbook.remove(self.template_sheet)

        return super().parse_template(expression_row)

    def set_course_information(self, course):
        pass

    def begin_section(self, section_name):
        title = sanitize_title(section_name)
        self.current_sheet = self.workbook.create_sheet(title=title, index=self.sheet_insert_index + self.section_count)
        self.current_range = CellRange(self.template_range.coord)
        self.section_count += 1  # insert sheets in ascending order
        WorksheetCopy(source_worksheet=self.template_sheet, target_worksheet=self.current_sheet).copy_worksheet()

    def end_section(self, section_name=None):
        # TODO: copy more properties of template_table
        # TODO: make sure that tables have at least one entry
        table = Table(ref=self.current_range.coord, displayName="DATA_{}".format(self.section_count))
        self.current_sheet.add_table(table)

    def write_row(self, row):
        row_iter = iter(row)
        self.current_range.expand(down=1)
        for c in self.current_range.bottom:
            self.current_sheet.cell(*c).value = next(row_iter)

    def get_data(self):
        with NamedTemporaryFile() as file:
            self.workbook.save(file.name)
            file.seek(0)
            stream = file.read()
        return stream




class ExcelZipWriter(ExcelWriter):
    """ The ExcelZipWriter begins a new .xlsx file for each new section.
    """

    mimetype = 'application/zip'
    extension = 'zip'

    def __init__(self, template):
        ExcelWriter.__init__(self, template)
        self.check_for_expressions()
        self.tempfile = NamedTemporaryFile()
        self.zip = ZipFile(self.tempfile, 'w')
        self.is_single_section = False
        self.single_section_data = None

    def set_is_single_section(self, single: bool):
        self.is_single_section = single

    def check_for_expressions(self):
        # set course information
        self.information_sheet = self.workbook.get_sheet_by_name("Notenliste")
        expressions = []
        self.coordinates = []
        max_row = self.information_sheet.max_row
        # iterate sheet to find jinja expressions
        # for row in self.information_sheet.iter_rows(min_row=30, min_col=1, max_row=max_row, max_col=3):
        for row in self.information_sheet.iter_rows(min_row=1, min_col=1, max_row=max_row, max_col=3):
            for cell in row:
                if cell.value is not None:
                    key = cell.value
                    options = [
                        'course.ger',
                        'course.ects_points',
                        'course.name',
                        'course.alternative',
                        'course.name_english',
                        'semester',
                        'exam_date',
                        'course.teacher_name'
                    ]
                    # in case of integers they need to be converted to a string
                    if type(key) is int:
                        key = str(key)
                    # if one of the strings is equal, it gets added to the information list
                    if any(key in word for word in options):
                        self.coordinates.append(cell.coordinate)
                        expressions.append(key)
                        cell.value = None

        # gets converted into callable expression
        self.course_information = [app.jinja_env.compile_expression(e) for e in expressions]

    def set_course_information(self, course):
        semester = app.config['SEMESTER_NAME_SHORT']
        exam_date = app.config['EXAM_DATE']
        # convert jinja expressions into writable expression with the required data
        expression_column = [cell_template(dict(course=course, semester=semester, exam_date=exam_date))
                             for cell_template in self.course_information]

        # write the information column starting at the first found cell
        for i in range(len(expression_column)):
            cell = self.information_sheet[self.coordinates[i]]
            cell.value = expression_column[i]

    def begin_section(self, section_name):
        # If this is not the first section, set is_single_section to False
        if self.is_single_section and self.section_count > 0:
            self.is_single_section = False
        # use title of template sheet
        super().begin_section(section_name=self.template_sheet.title)

    def end_section(self, section_name):
        super().end_section(section_name)
        with NamedTemporaryFile() as file:
            self.workbook.save(file.name)

            if self.is_single_section:
                # If it's a single section, store the data
                with open(file.name, 'rb') as single_file:
                    self.single_section_data = single_file.read()
            else:
                # Otherwise, write to the ZIP file
                self.zip.write(file.name, "{}.xlsx".format(section_name))
        # Restore template workbook to initial state
        self.workbook.remove(self.current_sheet)
        self.section_count -= 1

    def get_data(self):
        if self.is_single_section and self.single_section_data:
            # update mimetype and extension to export single xlsx file
            self.mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            self.extension = 'xlsx'
            return self.single_section_data
        else:
            self.zip.close()
            self.tempfile.seek(0)
            return self.tempfile.read()


class SingleSectionExcelWriter(ExcelWriter):
    section = None

    def set_course_information(self, course):
        pass

    def begin_section(self, section_name):
        if not self.section:
            super().begin_section(section_name)
        self.section = section_name

    def end_section(self, section_name):
        pass

    def get_data(self):
        if self.section:
            super().end_section(self.section)
        return super().get_data()

